"""
services/excel_reader.py
Lector de solicitudes desde archivo Excel.

Este módulo:
- Lee solicitudes pendientes de la hoja "Solicitudes"
- Valida los datos antes de procesarlos
- Resuelve IDs de pruebas a objetos CourseData
- Resuelve nombre de generalista a email
- Actualiza el estado después del procesamiento

Uso:
    from services.excel_reader import ExcelReader
    from services.course_cache import CourseCache
    
    cache = CourseCache()
    reader = ExcelReader("plantilla_solicitudes.xlsx", cache)
    
    # Obtener solicitudes pendientes
    solicitudes = reader.get_pending_requests()
    
    for solicitud in solicitudes:
        print(f"Candidato: {solicitud.nombres} {solicitud.apellidos}")
        print(f"Pruebas: {[p.name for p in solicitud.pruebas]}")
        print(f"Generalista: {solicitud.generalista_email}")
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Optional, List
from dataclasses import dataclass, field

from openpyxl import load_workbook

from core.logger import get_logger
from models.course import CourseData

logger = get_logger(__name__)


@dataclass
class SolicitudPrueba:
    """
    Representa una solicitud de asignación de pruebas.
    
    Contiene todos los datos necesarios para procesar la solicitud
    y enviar el correo con los resultados.
    """
    # Identificación
    codigo_solicitud: str
    row_number: int  # Fila en el Excel (para actualizar estado)
    
    # Datos del candidato
    nombres: str
    apellidos: str
    email: str
    telefono: str = ""
    ciudad: str = ""
    
    # Configuración
    duracion_dias: int = 1
    
    # Generalista
    generalista_nombre: str = ""
    generalista_email: str = ""
    
    # Pruebas (resueltas desde IDs)
    pruebas_ids: List[str] = field(default_factory=list)
    pruebas: List[CourseData] = field(default_factory=list)
    
    # Estado (para actualización post-proceso)
    estado: str = "Pendiente"
    fecha_procesado: Optional[datetime] = None
    observaciones: str = ""
    
    # Resultados del procesamiento
    usuario_moodle: str = ""
    password: str = ""
    documento_generado: str = ""
    
    @property
    def nombre_completo(self) -> str:
        """Nombre completo del candidato."""
        return f"{self.nombres} {self.apellidos}"
    
    @property
    def es_valida(self) -> bool:
        """Verifica si la solicitud tiene datos mínimos válidos."""
        return bool(
            self.nombres and
            self.apellidos and
            self.email and
            self.pruebas_ids
        )
    
    @property
    def identificador(self) -> str:
        """Identificador para mostrar (código o email)."""
        return self.codigo_solicitud if self.codigo_solicitud else self.email
    
    @property
    def errores_validacion(self) -> List[str]:
        """Lista de errores de validación."""
        errores = []
        if not self.nombres:
            errores.append("Falta nombres")
        if not self.apellidos:
            errores.append("Falta apellidos")
        if not self.email:
            errores.append("Falta email")
        if not self.pruebas_ids:
            errores.append("Falta IDs de pruebas")
        return errores


class ExcelReader:
    """
    Lee y procesa solicitudes desde archivo Excel.
    
    Funcionalidades:
    - Leer solicitudes pendientes
    - Validar datos
    - Resolver IDs de pruebas
    - Resolver generalistas
    - Actualizar estados después del procesamiento
    """
    
    # Mapeo de columnas (0-indexed desde columna A)
    COL_CODIGO = 0       # A: codigo_solicitud
    COL_NOMBRES = 1      # B: nombres
    COL_APELLIDOS = 2    # C: apellidos
    COL_EMAIL = 3        # D: email
    COL_TELEFONO = 4     # E: telefono
    COL_CIUDAD = 5       # F: ciudad
    COL_DURACION = 6     # G: duracion_dias
    COL_GENERALISTA = 7  # H: generalista
    COL_PRUEBAS = 8      # I: pruebas_ids
    COL_ESTADO = 9       # J: estado
    COL_FECHA = 10       # K: fecha_procesado
    COL_OBS = 11         # L: observaciones
    
    def __init__(self, filepath: str, cache=None):
        """
        Inicializa el lector.
        
        Args:
            filepath: Ruta al archivo Excel
            cache: Instancia de CourseCache
        """
        self.filepath = filepath
        self.cache = cache
        self._generalistas = {}  # Cache de nombre -> email
    
    def _get_cache(self):
        """Obtiene el cache, cargándolo si es necesario."""
        if self.cache is None:
            from services.course_cache import CourseCache
            self.cache = CourseCache()
        return self.cache
    
    def _load_generalistas(self, wb) -> dict:
        """
        Carga el mapeo de generalistas (nombre -> email).
        
        Args:
            wb: Workbook abierto
            
        Returns:
            Diccionario {nombre: email}
        """
        if self._generalistas:
            return self._generalistas
        
        if "Generalistas" not in wb.sheetnames:
            logger.warning("No se encontró hoja 'Generalistas'")
            return {}
        
        sheet = wb["Generalistas"]
        
        for row in range(2, sheet.max_row + 1):
            nombre = sheet.cell(row=row, column=2).value  # Columna B
            email = sheet.cell(row=row, column=3).value   # Columna C
            
            if nombre and email:
                self._generalistas[str(nombre).strip()] = str(email).strip()
        
        logger.debug(f"Cargadas {len(self._generalistas)} generalistas")
        return self._generalistas
    
    def _parse_pruebas_ids(self, value: str) -> List[str]:
        """
        Parsea la cadena de IDs de pruebas.
        
        Args:
            value: String con IDs separados por punto y coma (ej: "611;590;616")
            
        Returns:
            Lista de IDs como strings
        """
        if not value:
            return []
        
        # Limpiar y separar
        ids = []
        for part in str(value).split(";"):
            part = part.strip()
            if part:
                # Intentar convertir a entero y de vuelta a string para validar
                try:
                    ids.append(str(int(float(part))))
                except ValueError:
                    logger.warning(f"ID de prueba inválido: {part}")
        
        return ids
    
    def _resolve_pruebas(self, ids: List[str]) -> List[CourseData]:
        """
        Resuelve IDs de pruebas a objetos CourseData.
        
        Args:
            ids: Lista de IDs de pruebas
            
        Returns:
            Lista de CourseData encontrados
        """
        cache = self._get_cache()
        pruebas = []
        
        for course_id in ids:
            course = cache.find_by_id(course_id)
            if course:
                pruebas.append(course)
            else:
                logger.warning(f"Prueba no encontrada en cache: ID {course_id}")
        
        return pruebas
    
    def _read_row(self, sheet, row: int, generalistas: dict) -> Optional[SolicitudPrueba]:
        """
        Lee una fila y la convierte en SolicitudPrueba.
        
        Args:
            sheet: Hoja de Excel
            row: Número de fila
            generalistas: Diccionario de generalistas
            
        Returns:
            SolicitudPrueba o None si la fila está vacía
        """
        # Leer valores
        codigo = sheet.cell(row=row, column=self.COL_CODIGO + 1).value
        nombres = sheet.cell(row=row, column=self.COL_NOMBRES + 1).value
        apellidos = sheet.cell(row=row, column=self.COL_APELLIDOS + 1).value
        email = sheet.cell(row=row, column=self.COL_EMAIL + 1).value
        telefono = sheet.cell(row=row, column=self.COL_TELEFONO + 1).value
        ciudad = sheet.cell(row=row, column=self.COL_CIUDAD + 1).value
        duracion = sheet.cell(row=row, column=self.COL_DURACION + 1).value
        generalista = sheet.cell(row=row, column=self.COL_GENERALISTA + 1).value
        pruebas_str = sheet.cell(row=row, column=self.COL_PRUEBAS + 1).value
        estado = sheet.cell(row=row, column=self.COL_ESTADO + 1).value
        
        # Verificar si la fila tiene datos
        if not codigo and not email:
            return None
        
        # Parsear IDs de pruebas
        pruebas_ids = self._parse_pruebas_ids(pruebas_str)
        
        # Resolver pruebas desde cache
        pruebas = self._resolve_pruebas(pruebas_ids)
        
        # Resolver email de generalista
        generalista_nombre = str(generalista).strip() if generalista else ""
        generalista_email = generalistas.get(generalista_nombre, "")
        
        # Parsear duración
        try:
            duracion_dias = int(duracion) if duracion else 1
            duracion_dias = max(1, min(3, duracion_dias))  # Limitar entre 1 y 3
        except (ValueError, TypeError):
            duracion_dias = 1
        
        return SolicitudPrueba(
            codigo_solicitud=str(codigo).strip() if codigo else "",
            row_number=row,
            nombres=str(nombres).strip() if nombres else "",
            apellidos=str(apellidos).strip() if apellidos else "",
            email=str(email).strip().lower() if email else "",
            telefono=str(telefono).strip() if telefono else "",
            ciudad=str(ciudad).strip() if ciudad else "",
            duracion_dias=duracion_dias,
            generalista_nombre=generalista_nombre,
            generalista_email=generalista_email,
            pruebas_ids=pruebas_ids,
            pruebas=pruebas,
            estado=str(estado).strip() if estado else "Pendiente",
        )
    
    def get_pending_requests(self) -> List[SolicitudPrueba]:
        """
        Obtiene todas las solicitudes pendientes.
        
        Returns:
            Lista de SolicitudPrueba con estado "Pendiente"
        """
        if not os.path.exists(self.filepath):
            logger.error(f"Archivo no existe: {self.filepath}")
            return []
        
        logger.info(f"Leyendo solicitudes de: {self.filepath}")
        
        try:
            wb = load_workbook(self.filepath, data_only=True)
            
            if "Solicitudes" not in wb.sheetnames:
                logger.error("No se encontró hoja 'Solicitudes'")
                return []
            
            # Cargar generalistas
            generalistas = self._load_generalistas(wb)
            
            sheet = wb["Solicitudes"]
            solicitudes = []
            
            # Iterar filas (empezando en 2, después del encabezado)
            for row in range(2, sheet.max_row + 1):
                solicitud = self._read_row(sheet, row, generalistas)
                
                if solicitud is None:
                    continue  # Fila vacía
                
                # Solo incluir pendientes
                if solicitud.estado.lower() == "pendiente":
                    if solicitud.es_valida:
                        solicitudes.append(solicitud)
                        logger.debug(f"Solicitud válida: {solicitud.codigo_solicitud} - {solicitud.nombre_completo}")
                    else:
                        errores = ", ".join(solicitud.errores_validacion)
                        logger.warning(f"Solicitud inválida en fila {row}: {errores}")
            
            wb.close()
            
            logger.info(f"Encontradas {len(solicitudes)} solicitudes pendientes")
            return solicitudes
            
        except Exception as e:
            logger.error(f"Error leyendo Excel: {e}")
            return []
    
    def get_all_requests(self) -> List[SolicitudPrueba]:
        """
        Obtiene todas las solicitudes (sin importar estado).
        
        Returns:
            Lista de todas las SolicitudPrueba
        """
        if not os.path.exists(self.filepath):
            logger.error(f"Archivo no existe: {self.filepath}")
            return []
        
        try:
            wb = load_workbook(self.filepath, data_only=True)
            
            if "Solicitudes" not in wb.sheetnames:
                return []
            
            generalistas = self._load_generalistas(wb)
            sheet = wb["Solicitudes"]
            solicitudes = []
            
            for row in range(2, sheet.max_row + 1):
                solicitud = self._read_row(sheet, row, generalistas)
                if solicitud and solicitud.codigo_solicitud:
                    solicitudes.append(solicitud)
            
            wb.close()
            return solicitudes
            
        except Exception as e:
            logger.error(f"Error leyendo Excel: {e}")
            return []
    
    def update_request_status(
        self,
        row_number: int,
        estado: str,
        observaciones: str = "",
        usuario_moodle: str = "",
        password: str = "",
        documento: str = ""
    ) -> bool:
        """
        Actualiza el estado de una solicitud en el Excel.
        
        Args:
            row_number: Número de fila a actualizar
            estado: Nuevo estado ("Completado", "Error", etc.)
            observaciones: Observaciones del procesamiento
            usuario_moodle: Username asignado
            password: Contraseña generada
            documento: Ruta del documento generado
            
        Returns:
            True si se actualizó correctamente
        """
        try:
            wb = load_workbook(self.filepath)
            sheet = wb["Solicitudes"]
            
            # Actualizar estado
            sheet.cell(row=row_number, column=self.COL_ESTADO + 1, value=estado)
            
            # Actualizar fecha
            sheet.cell(row=row_number, column=self.COL_FECHA + 1, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
            
            # Actualizar observaciones
            if observaciones:
                sheet.cell(row=row_number, column=self.COL_OBS + 1, value=observaciones)
            
            wb.save(self.filepath)
            wb.close()
            
            logger.debug(f"Estado actualizado en fila {row_number}: {estado}")
            return True
            
        except Exception as e:
            logger.error(f"Error actualizando estado: {e}")
            return False
    
    def add_to_processed(
        self,
        solicitud: SolicitudPrueba,
        usuario_moodle: str,
        password: str,
        pruebas_asignadas: List[str],
        documento: str,
        estado: str = "Completado"
    ) -> bool:
        """
        Agrega un registro a la hoja de Procesados.
        
        Args:
            solicitud: Solicitud procesada
            usuario_moodle: Username asignado
            password: Contraseña
            pruebas_asignadas: Lista de nombres de pruebas asignadas
            documento: Ruta del documento generado
            estado: Estado final
            
        Returns:
            True si se agregó correctamente
        """
        try:
            wb = load_workbook(self.filepath)
            
            if "Procesados" not in wb.sheetnames:
                logger.warning("No se encontró hoja 'Procesados'")
                wb.close()
                return False
            
            sheet = wb["Procesados"]
            
            # Encontrar siguiente fila vacía
            next_row = sheet.max_row + 1
            if next_row == 2 and not sheet.cell(row=2, column=1).value:
                next_row = 2
            
            # Escribir datos
            sheet.cell(row=next_row, column=1, value=solicitud.codigo_solicitud)
            sheet.cell(row=next_row, column=2, value=solicitud.email)
            sheet.cell(row=next_row, column=3, value=usuario_moodle)
            sheet.cell(row=next_row, column=4, value=password)
            sheet.cell(row=next_row, column=5, value="; ".join(pruebas_asignadas))
            sheet.cell(row=next_row, column=6, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
            sheet.cell(row=next_row, column=7, value=estado)
            sheet.cell(row=next_row, column=8, value=documento)
            sheet.cell(row=next_row, column=9, value=solicitud.generalista_email)
            
            wb.save(self.filepath)
            wb.close()
            
            logger.info(f"Registro agregado a Procesados: {solicitud.codigo_solicitud}")
            return True
            
        except Exception as e:
            logger.error(f"Error agregando a Procesados: {e}")
            return False


# ==================== Función de conveniencia ====================

def load_pending_requests(filepath: str, cache=None) -> List[SolicitudPrueba]:
    """
    Función de conveniencia para cargar solicitudes pendientes.
    
    Args:
        filepath: Ruta al archivo Excel
        cache: Cache de cursos (opcional)
        
    Returns:
        Lista de solicitudes pendientes
    """
    reader = ExcelReader(filepath, cache)
    return reader.get_pending_requests()