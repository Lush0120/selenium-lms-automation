"""
utils/excel_template.py
Generador y actualizador de plantilla Excel para solicitudes.

Este módulo:
- Crea la plantilla inicial con todas las hojas
- Actualiza la hoja de catálogo de pruebas desde el cache
- Configura validaciones de datos (listas desplegables)
- Genera fórmula de previsualización de pruebas

Uso:
    from utils.excel_template import ExcelTemplateGenerator
    from services.course_cache import CourseCache
    
    cache = CourseCache()
    generator = ExcelTemplateGenerator(cache)
    
    # Crear plantilla nueva
    generator.create_template("plantilla_solicitudes.xlsx")
    
    # Actualizar solo catálogo de pruebas
    generator.update_catalog("plantilla_solicitudes.xlsx")
"""

import os
from pathlib import Path
from datetime import datetime
from typing import Optional, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

from core.logger import get_logger

logger = get_logger(__name__)


# Colores para el Excel
COLORS = {
    "header_bg": "4472C4",      # Azul oscuro
    "header_font": "FFFFFF",    # Blanco
    "alt_row": "D9E2F3",        # Azul claro
    "success": "C6EFCE",        # Verde claro
    "warning": "FFEB9C",        # Amarillo claro
    "error": "FFC7CE",          # Rojo claro
    "auto_fill": "E2EFDA",      # Verde muy claro (campos auto)
}

# Estilos
HEADER_FONT = Font(bold=True, color=COLORS["header_font"], size=11)
HEADER_FILL = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


class ExcelTemplateGenerator:
    """
    Genera y actualiza plantillas Excel para solicitudes de pruebas.
    
    Estructura del Excel:
    - Hoja "Solicitudes": Datos de candidatos y pruebas a asignar
    - Hoja "Generalistas": Catálogo de generalistas con emails
    - Hoja "Catalogo_Pruebas": Lista de pruebas disponibles (desde cache)
    - Hoja "Procesados": Registro de solicitudes procesadas (auto)
    """
    
    # Columnas de la hoja Solicitudes
    SOLICITUDES_COLUMNS = [
        ("codigo_solicitud", "Código Solicitud", 18),
        ("nombres", "Nombres", 25),
        ("apellidos", "Apellidos", 25),
        ("email", "Email Candidato", 30),
        ("telefono", "Teléfono", 15),
        ("ciudad", "Ciudad", 15),
        ("duracion_dias", "Días", 8),
        ("generalista", "Generalista", 25),
        ("pruebas_ids", "IDs Pruebas (;)", 25),
        ("estado", "Estado", 12),
        ("fecha_procesado", "Fecha Procesado", 18),
        ("observaciones", "Observaciones", 40),
    ]
    
    # Columnas de la hoja Generalistas
    GENERALISTAS_COLUMNS = [
        ("id", "ID", 8),
        ("nombre", "Nombre", 30),
        ("email", "Email", 35),
    ]
    
    # Columnas de la hoja Catalogo_Pruebas
    CATALOGO_COLUMNS = [
        ("course_id", "ID", 8),
        ("categoria", "Categoría", 15),
        ("nombre", "Nombre de la Prueba", 55),
    ]
    
    # Columnas de la hoja Procesados
    PROCESADOS_COLUMNS = [
        ("codigo_solicitud", "Código Solicitud", 18),
        ("email", "Email", 30),
        ("usuario_moodle", "Usuario Moodle", 15),
        ("password", "Contraseña", 15),
        ("pruebas_asignadas", "Pruebas Asignadas", 40),
        ("fecha_procesado", "Fecha Procesado", 18),
        ("estado", "Estado", 12),
        ("documento", "Documento Generado", 50),
        ("generalista_email", "Email Generalista", 30),
    ]
    
    def __init__(self, cache=None):
        """
        Inicializa el generador.
        
        Args:
            cache: Instancia de CourseCache (opcional, se carga si no se proporciona)
        """
        self.cache = cache
    
    def _get_cache(self):
        """Obtiene el cache, cargándolo si es necesario."""
        if self.cache is None:
            from services.course_cache import CourseCache
            self.cache = CourseCache()
        return self.cache
    
    def _apply_header_style(self, sheet, row: int, columns: list) -> None:
        """Aplica estilo a la fila de encabezados."""
        for col_idx, (_, header, width) in enumerate(columns, 1):
            cell = sheet.cell(row=row, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER
            sheet.column_dimensions[get_column_letter(col_idx)].width = width
    
    def _create_solicitudes_sheet(self, wb: Workbook) -> None:
        """Crea la hoja de Solicitudes."""
        if "Solicitudes" in wb.sheetnames:
            del wb["Solicitudes"]
        
        sheet = wb.create_sheet("Solicitudes", 0)
        
        # Encabezados
        self._apply_header_style(sheet, 1, self.SOLICITUDES_COLUMNS)
        
        # Congelar primera fila
        sheet.freeze_panes = "A2"
        
        # Agregar filas de ejemplo vacías con formato
        for row in range(2, 102):  # 100 filas para datos
            for col_idx in range(1, len(self.SOLICITUDES_COLUMNS) + 1):
                cell = sheet.cell(row=row, column=col_idx)
                cell.border = THIN_BORDER
                
                # Colorear columnas auto-generadas
                col_name = self.SOLICITUDES_COLUMNS[col_idx - 1][0]
                if col_name in ["pruebas_preview", "estado", "fecha_procesado", "observaciones"]:
                    cell.fill = PatternFill(start_color=COLORS["auto_fill"], end_color=COLORS["auto_fill"], fill_type="solid")
        
        # Valor por defecto para estado
        for row in range(2, 102):
            sheet.cell(row=row, column=11, value="Pendiente")
        
        logger.debug("Hoja 'Solicitudes' creada")
    
    def _create_generalistas_sheet(self, wb: Workbook) -> None:
        """Crea la hoja de Generalistas."""
        if "Generalistas" in wb.sheetnames:
            del wb["Generalistas"]
        
        sheet = wb.create_sheet("Generalistas")
        
        # Encabezados
        self._apply_header_style(sheet, 1, self.GENERALISTAS_COLUMNS)
        
        # Congelar primera fila
        sheet.freeze_panes = "A2"
        
        # Generalistas actuales de Q-Vision
        generalistas_qvision = [
            (1, "Carolina Hernandez Vaenas", "phernandez@qvision.us"),
            (2, "Paulina Morales Echeverry", "mmorales@qvision.us"),
            (3, "Luna Garcia Martinez", "mgarcia@qvision.us"),
            (4, "Lady Poveda Clavijo", "ljpoveda@qvision.us"),
            (5, "Sara Estefania Orozco Montoya", "sorozco@qvision.us"),
        ]
        
        for row_idx, (id_, nombre, email) in enumerate(generalistas_qvision, 2):
            sheet.cell(row=row_idx, column=1, value=id_).border = THIN_BORDER
            sheet.cell(row=row_idx, column=2, value=nombre).border = THIN_BORDER
            sheet.cell(row=row_idx, column=3, value=email).border = THIN_BORDER
        
        logger.debug("Hoja 'Generalistas' creada")
    
    def _create_catalogo_sheet(self, wb: Workbook) -> None:
        """Crea la hoja de Catálogo de Pruebas."""
        if "Catalogo_Pruebas" in wb.sheetnames:
            del wb["Catalogo_Pruebas"]
        
        sheet = wb.create_sheet("Catalogo_Pruebas")
        
        # Encabezados
        self._apply_header_style(sheet, 1, self.CATALOGO_COLUMNS)
        
        # Congelar primera fila
        sheet.freeze_panes = "A2"
        
        # Cargar pruebas desde cache
        cache = self._get_cache()
        courses = cache.get_all()
        
        if not courses:
            logger.warning("Cache vacío, no se agregaron pruebas al catálogo")
            # Agregar mensaje de aviso
            sheet.cell(row=2, column=1, value="⚠️ Sincroniza el catálogo para ver las pruebas disponibles")
            sheet.merge_cells("A2:C2")
            return
        
        # Importar enum para comparación
        from models.course import CourseCategory
        
        # Ordenar por categoría y nombre
        courses_sorted = sorted(courses, key=lambda c: (c.category.value if c.category else "", c.name))
        
        for row_idx, course in enumerate(courses_sorted, 2):
            # Determinar categoría usando el enum
            if course.category == CourseCategory.PRUEBAS_DIAGNOSTICAS:
                categoria = "Diagnóstica"
            elif course.category == CourseCategory.PRUEBAS_TECNICAS:
                categoria = "Técnica"
            else:
                categoria = "Otra"
            
            cell_id = sheet.cell(row=row_idx, column=1, value=int(course.course_id))
            cell_cat = sheet.cell(row=row_idx, column=2, value=categoria)
            cell_name = sheet.cell(row=row_idx, column=3, value=course.name)
            
            cell_id.border = THIN_BORDER
            cell_cat.border = THIN_BORDER
            cell_name.border = THIN_BORDER
            
            # Colorear filas alternadas
            if row_idx % 2 == 0:
                fill = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")
                cell_id.fill = fill
                cell_cat.fill = fill
                cell_name.fill = fill
        
        logger.info(f"Catálogo actualizado con {len(courses)} pruebas")
    
    def _create_procesados_sheet(self, wb: Workbook) -> None:
        """Crea la hoja de Procesados."""
        if "Procesados" in wb.sheetnames:
            del wb["Procesados"]
        
        sheet = wb.create_sheet("Procesados")
        
        # Encabezados
        self._apply_header_style(sheet, 1, self.PROCESADOS_COLUMNS)
        
        # Congelar primera fila
        sheet.freeze_panes = "A2"
        
        logger.debug("Hoja 'Procesados' creada")
    
    def _setup_data_validations(self, wb: Workbook) -> None:
        """Configura las validaciones de datos (listas desplegables)."""
        sheet = wb["Solicitudes"]
        
        # Lista desplegable para Generalistas (columna H)
        # Referencia a la columna "nombre" de la hoja Generalistas
        dv_generalistas = DataValidation(
            type="list",
            formula1="Generalistas!$B$2:$B$100",
            allow_blank=True,
            showDropDown=False
        )
        dv_generalistas.error = "Selecciona una generalista de la lista"
        dv_generalistas.errorTitle = "Generalista inválida"
        dv_generalistas.prompt = "Selecciona la generalista"
        dv_generalistas.promptTitle = "Generalista"
        sheet.add_data_validation(dv_generalistas)
        dv_generalistas.add("H2:H101")
        
        # Lista desplegable para Duración (columna G)
        dv_duracion = DataValidation(
            type="list",
            formula1='"1,2,3"',
            allow_blank=False,
            showDropDown=False
        )
        dv_duracion.error = "Ingresa 1, 2 o 3 días"
        dv_duracion.errorTitle = "Duración inválida"
        sheet.add_data_validation(dv_duracion)
        dv_duracion.add("G2:G101")
        
        # Lista desplegable para Estado (columna J - sin columna preview)
        dv_estado = DataValidation(
            type="list",
            formula1='"Pendiente,Procesando,Completado,Error"',
            allow_blank=False,
            showDropDown=False
        )
        sheet.add_data_validation(dv_estado)
        dv_estado.add("J2:J101")
        
        logger.debug("Validaciones de datos configuradas")
    
    def _setup_conditional_formatting(self, wb: Workbook) -> None:
        """Configura formato condicional para estados."""
        sheet = wb["Solicitudes"]
        
        # Estado ahora está en columna J (índice 10)
        # Verde para Completado
        green_fill = PatternFill(start_color=COLORS["success"], end_color=COLORS["success"], fill_type="solid")
        rule_completado = FormulaRule(
            formula=['$J2="Completado"'],
            fill=green_fill
        )
        sheet.conditional_formatting.add("A2:L101", rule_completado)
        
        # Amarillo para Procesando
        yellow_fill = PatternFill(start_color=COLORS["warning"], end_color=COLORS["warning"], fill_type="solid")
        rule_procesando = FormulaRule(
            formula=['$J2="Procesando"'],
            fill=yellow_fill
        )
        sheet.conditional_formatting.add("A2:L101", rule_procesando)
        
        # Rojo para Error
        red_fill = PatternFill(start_color=COLORS["error"], end_color=COLORS["error"], fill_type="solid")
        rule_error = FormulaRule(
            formula=['$J2="Error"'],
            fill=red_fill
        )
        sheet.conditional_formatting.add("A2:L101", rule_error)
        
        logger.debug("Formato condicional configurado")
    
    def create_template(self, filepath: str) -> str:
        """
        Crea una plantilla Excel nueva.
        
        Args:
            filepath: Ruta donde guardar el archivo
            
        Returns:
            Ruta del archivo creado
        """
        logger.info(f"Creando plantilla Excel: {filepath}")
        
        wb = Workbook()
        
        # Eliminar hoja por defecto
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Crear hojas
        self._create_solicitudes_sheet(wb)
        self._create_generalistas_sheet(wb)
        self._create_catalogo_sheet(wb)
        self._create_procesados_sheet(wb)
        
        # Configurar validaciones
        self._setup_data_validations(wb)
        self._setup_conditional_formatting(wb)
        
        # Activar hoja de solicitudes
        wb.active = wb["Solicitudes"]
        
        # Guardar
        wb.save(filepath)
        
        logger.info(f"Plantilla creada exitosamente: {filepath}")
        return filepath
    
    def update_catalog(self, filepath: str) -> bool:
        """
        Actualiza solo la hoja de catálogo de pruebas en un Excel existente.
        
        Args:
            filepath: Ruta del archivo Excel
            
        Returns:
            True si se actualizó correctamente
        """
        if not os.path.exists(filepath):
            logger.error(f"Archivo no existe: {filepath}")
            return False
        
        logger.info(f"Actualizando catálogo en: {filepath}")
        
        try:
            wb = load_workbook(filepath)
            self._create_catalogo_sheet(wb)
            wb.save(filepath)
            logger.info("Catálogo actualizado")
            return True
        except Exception as e:
            logger.error(f"Error actualizando catálogo: {e}")
            return False
    
    def sync_from_cache(self, filepath: str) -> bool:
        """
        Sincroniza el Excel con el cache actual.
        
        Si el archivo no existe, lo crea.
        Si existe, actualiza solo el catálogo de pruebas.
        
        Args:
            filepath: Ruta del archivo Excel
            
        Returns:
            True si se sincronizó correctamente
        """
        if os.path.exists(filepath):
            return self.update_catalog(filepath)
        else:
            self.create_template(filepath)
            return True


# ==================== Función de conveniencia ====================

def generate_template(filepath: str = "plantilla_solicitudes.xlsx", cache=None) -> str:
    """
    Función de conveniencia para generar una plantilla.
    
    Args:
        filepath: Ruta del archivo
        cache: Cache de cursos (opcional)
        
    Returns:
        Ruta del archivo creado
    """
    generator = ExcelTemplateGenerator(cache)
    return generator.create_template(filepath)


def sync_template(filepath: str = "plantilla_solicitudes.xlsx", cache=None) -> bool:
    """
    Función de conveniencia para sincronizar plantilla con cache.
    
    Args:
        filepath: Ruta del archivo
        cache: Cache de cursos (opcional)
        
    Returns:
        True si se sincronizó
    """
    generator = ExcelTemplateGenerator(cache)
    return generator.sync_from_cache(filepath)


# ==================== Gestión de Generalistas ====================

class GeneralistasManager:
    """
    Gestiona el listado de generalistas en el archivo Excel.
    
    Permite agregar, eliminar y listar generalistas sin modificar código.
    Soporta búsqueda por ID, nombre o email.
    """
    
    def __init__(self, filepath: str):
        """
        Inicializa el gestor.
        
        Args:
            filepath: Ruta al archivo Excel
            
        Raises:
            ExcelFileNotFoundError: Si el archivo no existe
        """
        self.filepath = filepath
        self._validate_file()
    
    def _validate_file(self) -> None:
        """Valida que el archivo Excel exista."""
        if not os.path.exists(self.filepath):
            from core.exceptions import ExcelFileNotFoundError
            raise ExcelFileNotFoundError(self.filepath)
    
    def _validate_sheet(self, wb) -> None:
        """Valida que exista la hoja de Generalistas."""
        if "Generalistas" not in wb.sheetnames:
            from core.exceptions import ExcelSheetNotFoundError
            raise ExcelSheetNotFoundError("Generalistas", self.filepath)
    
    def _find_generalista_row(self, sheet, identifier: str) -> Optional[int]:
        """
        Busca un generalista por ID, nombre o email.
        
        Args:
            sheet: Hoja de Excel
            identifier: ID (número), nombre o email del generalista
            
        Returns:
            Número de fila o None si no se encuentra
        """
        identifier_str = str(identifier).strip()
        identifier_lower = identifier_str.lower()
        
        # Verificar si es un ID numérico
        try:
            id_num = int(identifier_str)
            is_numeric = True
        except ValueError:
            is_numeric = False
            id_num = None
        
        for row in range(2, sheet.max_row + 1):
            id_cell = sheet.cell(row=row, column=1).value
            nombre = sheet.cell(row=row, column=2).value
            email = sheet.cell(row=row, column=3).value
            
            if not nombre:
                continue
            
            # Buscar por ID numérico
            if is_numeric and id_cell:
                try:
                    if int(id_cell) == id_num:
                        return row
                except (ValueError, TypeError):
                    pass
            
            # Buscar por nombre o email
            if nombre and identifier_lower == str(nombre).lower().strip():
                return row
            if email and identifier_lower == str(email).lower().strip():
                return row
        
        return None
    
    def list_generalistas(self) -> List[dict]:
        """
        Lista todos los generalistas del Excel.
        
        Returns:
            Lista de diccionarios con {id, nombre, email, row}
            
        Raises:
            ExcelFileNotFoundError: Si el archivo no existe
            ExcelSheetNotFoundError: Si no existe la hoja
        """
        self._validate_file()
        
        try:
            wb = load_workbook(self.filepath)
            self._validate_sheet(wb)
            
            sheet = wb["Generalistas"]
            generalistas = []
            
            for row in range(2, sheet.max_row + 1):
                id_ = sheet.cell(row=row, column=1).value
                nombre = sheet.cell(row=row, column=2).value
                email = sheet.cell(row=row, column=3).value
                
                if nombre and email:
                    generalistas.append({
                        "id": int(id_) if id_ else row - 1,
                        "nombre": str(nombre).strip(),
                        "email": str(email).strip(),
                        "row": row
                    })
            
            wb.close()
            return generalistas
            
        except Exception as e:
            logger.error(f"Error listando generalistas: {e}")
            raise
    
    def get_generalista(self, identifier: str) -> Optional[dict]:
        """
        Obtiene un generalista por ID, nombre o email.
        
        Args:
            identifier: ID, nombre o email del generalista
            
        Returns:
            Diccionario con datos del generalista o None
        """
        self._validate_file()
        
        try:
            wb = load_workbook(self.filepath)
            self._validate_sheet(wb)
            
            sheet = wb["Generalistas"]
            row = self._find_generalista_row(sheet, identifier)
            
            if row is None:
                wb.close()
                return None
            
            generalista = {
                "id": int(sheet.cell(row=row, column=1).value or row - 1),
                "nombre": str(sheet.cell(row=row, column=2).value).strip(),
                "email": str(sheet.cell(row=row, column=3).value).strip(),
                "row": row
            }
            
            wb.close()
            return generalista
            
        except Exception as e:
            logger.error(f"Error obteniendo generalista: {e}")
            return None
    
    def add_generalista(self, nombre: str, email: str) -> bool:
        """
        Agrega un nuevo generalista al Excel.
        
        Args:
            nombre: Nombre completo del generalista
            email: Email del generalista
            
        Returns:
            True si se agregó correctamente
            
        Raises:
            InvalidGeneralistaDataError: Si los datos son inválidos
            GeneralistaAlreadyExistsError: Si ya existe
        """
        from core.exceptions import InvalidGeneralistaDataError, GeneralistaAlreadyExistsError
        
        # Validar datos
        if not nombre or not nombre.strip():
            raise InvalidGeneralistaDataError("El nombre es requerido")
        if not email or not email.strip():
            raise InvalidGeneralistaDataError("El email es requerido")
        if "@" not in email:
            raise InvalidGeneralistaDataError("El email no es válido")
        
        self._validate_file()
        
        try:
            wb = load_workbook(self.filepath)
            self._validate_sheet(wb)
            
            sheet = wb["Generalistas"]
            
            # Verificar si ya existe
            existing_row = self._find_generalista_row(sheet, email.strip())
            if existing_row:
                wb.close()
                raise GeneralistaAlreadyExistsError(email)
            
            # Encontrar siguiente fila vacía y siguiente ID
            next_row = 2
            max_id = 0
            
            for row in range(2, sheet.max_row + 2):
                id_cell = sheet.cell(row=row, column=1).value
                nombre_cell = sheet.cell(row=row, column=2).value
                
                if id_cell:
                    try:
                        max_id = max(max_id, int(id_cell))
                    except (ValueError, TypeError):
                        pass
                
                if not nombre_cell:
                    next_row = row
                    break
                else:
                    next_row = row + 1
            
            # Agregar nuevo generalista
            new_id = max_id + 1
            sheet.cell(row=next_row, column=1, value=new_id).border = THIN_BORDER
            sheet.cell(row=next_row, column=2, value=nombre.strip()).border = THIN_BORDER
            sheet.cell(row=next_row, column=3, value=email.strip().lower()).border = THIN_BORDER
            
            wb.save(self.filepath)
            wb.close()
            
            logger.info(f"Generalista agregado: {nombre} ({email})")
            return True
            
        except (InvalidGeneralistaDataError, GeneralistaAlreadyExistsError):
            raise
        except Exception as e:
            logger.error(f"Error agregando generalista: {e}")
            return False
    
    def remove_generalista(self, identifier: str) -> bool:
        """
        Elimina un generalista por ID, nombre o email.
        
        Args:
            identifier: ID (número), nombre o email del generalista
            
        Returns:
            True si se eliminó correctamente
            
        Raises:
            GeneralistaNotFoundError: Si no se encuentra el generalista
        """
        from core.exceptions import GeneralistaNotFoundError
        
        self._validate_file()
        
        try:
            wb = load_workbook(self.filepath)
            self._validate_sheet(wb)
            
            sheet = wb["Generalistas"]
            row_to_delete = self._find_generalista_row(sheet, identifier)
            
            if row_to_delete is None:
                wb.close()
                raise GeneralistaNotFoundError(identifier)
            
            # Guardar nombre para el log
            nombre_eliminado = sheet.cell(row=row_to_delete, column=2).value
            
            # Eliminar la fila
            sheet.delete_rows(row_to_delete)
            
            # Renumerar IDs
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row=row, column=1, value=row - 1)
            
            wb.save(self.filepath)
            wb.close()
            
            logger.info(f"Generalista eliminado: {nombre_eliminado} (identificador: {identifier})")
            return True
            
        except GeneralistaNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Error eliminando generalista: {e}")
            return False
    
    def update_generalista(self, identifier: str, nuevo_nombre: str = None, nuevo_email: str = None) -> bool:
        """
        Actualiza los datos de un generalista.
        
        Args:
            identifier: ID (número), nombre o email actual del generalista
            nuevo_nombre: Nuevo nombre (opcional)
            nuevo_email: Nuevo email (opcional)
            
        Returns:
            True si se actualizó correctamente
            
        Raises:
            GeneralistaNotFoundError: Si no se encuentra el generalista
            InvalidGeneralistaDataError: Si no se proporcionan datos
        """
        from core.exceptions import GeneralistaNotFoundError, InvalidGeneralistaDataError
        
        if not nuevo_nombre and not nuevo_email:
            raise InvalidGeneralistaDataError("Debe proporcionar al menos un dato para actualizar")
        
        self._validate_file()
        
        try:
            wb = load_workbook(self.filepath)
            self._validate_sheet(wb)
            
            sheet = wb["Generalistas"]
            row_to_update = self._find_generalista_row(sheet, identifier)
            
            if row_to_update is None:
                wb.close()
                raise GeneralistaNotFoundError(identifier)
            
            # Actualizar datos
            if nuevo_nombre:
                sheet.cell(row=row_to_update, column=2, value=nuevo_nombre.strip())
            if nuevo_email:
                sheet.cell(row=row_to_update, column=3, value=nuevo_email.strip().lower())
            
            wb.save(self.filepath)
            wb.close()
            
            logger.info(f"Generalista actualizado: {identifier}")
            return True
            
        except GeneralistaNotFoundError:
            raise
        except Exception as e:
            logger.error(f"Error actualizando generalista: {e}")
            return False


# ==================== Funciones de conveniencia para generalistas ====================

def list_generalistas(filepath: str = "plantilla_solicitudes.xlsx") -> List[dict]:
    """Lista todos los generalistas."""
    manager = GeneralistasManager(filepath)
    return manager.list_generalistas()


def add_generalista(filepath: str, nombre: str, email: str) -> bool:
    """Agrega un nuevo generalista."""
    manager = GeneralistasManager(filepath)
    return manager.add_generalista(nombre, email)


def remove_generalista(filepath: str, identifier: str) -> bool:
    """Elimina un generalista por nombre o email."""
    manager = GeneralistasManager(filepath)
    return manager.remove_generalista(identifier)


def update_generalista(filepath: str, identifier: str, nuevo_nombre: str = None, nuevo_email: str = None) -> bool:
    """Actualiza un generalista."""
    manager = GeneralistasManager(filepath)
    return manager.update_generalista(identifier, nuevo_nombre, nuevo_email)